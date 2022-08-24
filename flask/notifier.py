"""
Copyright: Copyright (C) 2022 2DEVIATE Inc. - All Rights Reserved
Product: Standard
Description: Notifier Script to query exchange data and send notifications (email)
Notifyees: craig@2deviate.com
Category: None

"""
import re
import logging
import os
import json
import shutil
import pandas
import urllib3
import time
import argparse
import tempfile
import asyncio
import datetime
from run import create_app
from app.db import db
from config import configs
import app.constants as const
from dateutil import parser
from functools import reduce

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email_validator import validate_email, EmailNotValidError

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Font

from xlsxwriter.utility import xl_cell_to_rowcol


logger = logging.getLogger(__name__)
dir_path = os.path.dirname(os.path.realpath(__file__))

logging.basicConfig(
    filename="/".join([dir_path, "notifier.log"]),
    level=logging.DEBUG,
    format="%(asctime)s:%(levelname)s:%(message)s",
    force=True,
)

# script globals
docker_db_name = None
docker_server_name = None
docker_server_port = None
docker_proxy_name = None

email_from_address = None
email_cc_address = None
email_bcc_address = None
email_attachment = None
email_subject = None

smtp_user = None
smtp_password = None
smtp_host = None
smtp_port = None

email_template_text = None
email_template_html = None
email_template_schema = None

class ExcelFormatter(object):
    """Simple Excel Formatter Class"""

    def __init__(self, workbook, worksheet) -> None:
        self.workbook = workbook
        self.worksheet = worksheet

    @property
    def rows(self):
        return self.worksheet.dim_rowmax

    @property
    def columns(self):
        return self.worksheet.dim_colmax

    def set_border(self, range, border):
        formatter = self.workbook.add_format({'border': border})
        self.worksheet.conditional_format(range, {'type': 'no_errors', 'format': formatter})
    
    def set_background_color(self, range, color):
        formatter = self.workbook.add_format()
        formatter.set_pattern(1)
        formatter.set_bg_color(color)
        self.worksheet.conditional_format(range, {'type': 'text', 'criteria': 'containing', 'value': '', 'format': formatter})
    
    def set_width(self, range, width):
        self.worksheet.set_column(range, width)
    
    def set_zoom(self, zoom):
        self.worksheet.set_zoom(zoom)

    def set_format(self, range, format):
        formatter = self.workbook.add_format({'num_format': format})
        self.worksheet.set_column(range, None, formatter)

    def set_condition(self, range, criteria, format):
        formatter = self.workbook.add_format(format)
        condition = criteria | {'format': formatter}
        self.worksheet.conditional_format(range, condition)        
        
    @staticmethod
    def format(formatter):
        dt0 = datetime.date.today()
        dt1 = dt0 + datetime.timedelta(days=60)
        dt2 = datetime.date.today()+datetime.timedelta(days=90)        
        formatter.set_condition(f'F5:F{5+formatter.rows}', {'type': 'date', 'criteria': 'between', 'minimum': dt0, 'maximum': dt1}, {'bg_color': '#FFC7CE'})
        formatter.set_condition(f'F5:F{5+formatter.rows}', {'type': 'date', 'criteria': 'between', 'minimum': dt1, 'maximum': dt2}, {'bg_color': '#FFEB9C'})
        formatter.set_border('H1:H1', 0)
        formatter.set_border('H2:H2', 0)
        formatter.set_border('H3:H3', 0)
        formatter.set_width('A:A', 15)
        formatter.set_width('B:B', 20)
        formatter.set_width('C:C', 22)
        formatter.set_width('D:D', 20)
        formatter.set_width('E:E', 25)
        formatter.set_width('F:F', 25)
        formatter.set_width('G:G', 25)
        formatter.set_width('H:H', 18)
        formatter.set_width('I:AI', 7)
        formatter.set_background_color('I1:I1', '#E2EFDA')
        formatter.set_background_color('R1:R1', '#C6E0B4')
        formatter.set_background_color('AA1:AA1', '#A9D08E')        
        formatter.set_zoom(75)

def get_data(method, url=None, headers=None, body=None, **attrs):
    result = None
    try:
        http = urllib3.PoolManager()
        if method == "POST":
            payload = payload
            encoded = urllib3.request.urlencode(payload)
            body = encoded.encode("ascii")
            request = http.request(method=method, url=url, body=body, headers=headers)
        elif method == "GET":
            request = http.request(method=method, url=url, headers=headers)
        return request.data
    except Exception as err:
        logger.error(err, exc_info=err)
    return result


def send_mail(from_addr, to_addrs, cc_addrs, bcc_addrs, df):
    "Sends mail with excel attachment"
    
    logger.info(f"sending mail, {from_addr=}, {to_addrs=}, {cc_addrs=}, {bcc_addrs=}")
    
    text_subject = email_template_text.strip() if email_template_text else None        
    text_part = MIMEText(text_subject, "plain")    
    
    msg_alternative = MIMEMultipart("alternative")
    msg_alternative.attach(text_part)    

    msg_mixed = MIMEMultipart("mixed")
    msg_mixed.attach(msg_alternative)

    if not df.empty:
        tmpdir = tempfile.mkdtemp()
        tmpfil = os.path.join(tmpdir,"template.xlsx")
        logger.info(f"created tmp file {tmpfil=}")
        template = os.path.join(dir_path, f"app/data/downloads/template.xlsx")        
        try:
            shutil.copyfile(template, tmpfil)
        except IOError as err:
            logger.error(f"Unable to copy tmp file {tmpfil=}", err)
            raise    
        
        writer = pandas.ExcelWriter(tmpfil, engine='xlsxwriter', datetime_format='mm/dd/yyyy', date_format='mm/dd/yyyy')        
        # write df
        df.to_excel(writer, sheet_name='Sales Planner', merge_cells=True)
        # get excel workbook and sheet
        workbook = writer.book
        worksheet = writer.sheets['Sales Planner']
        # format sheet
        formatter = ExcelFormatter(workbook, worksheet)
        ExcelFormatter.format(formatter)
        # save excel
        writer.save()
        
        # open binary attachment
        fp = open(tmpfil, "rb")
        attachment = MIMEApplication(fp.read(), _subtype="xls")
        fp.close()
        # clean up
        if os.path.exists(tmpfil):
            try:
                os.remove(tmpfil)  # remove tmp file
                shutil.rmtree(tmpdir)  # remove tmp folder
            except OSError as err:
                logger.error(f"Unable to remove file {tmpfil=} or directory {tmpdir=}", err)
        # attach header
        attachment.add_header("Content-Disposition", "attachment", filename=email_attachment)    
        msg_mixed.attach(attachment)
    # setup mail address's
    to_addrs = [addrs for addrs in set(to_addrs) if validate_addrs(addrs) is not None]
    cc_addrs = [addrs for addrs in set(cc_addrs) if validate_addrs(addrs) is not None]
    bcc_addrs = [addrs for addrs in set(bcc_addrs) if validate_addrs(addrs) is not None]
    # setup mime message
    msg_mixed["From"] = from_addr
    msg_mixed["To"] = ",".join(to_addrs)
    msg_mixed["Cc"] = ",".join(cc_addrs)
    msg_mixed["Bcc"] = ",".join(bcc_addrs)
    msg_mixed["Subject"] = email_subject    

    try:
        logger.info(f"login to smpt sever {smtp_host=}, {smtp_port=}") 
        smtp_obj = smtplib.SMTP_SSL(host=smtp_host, port=smtp_port)
        smtp_obj.ehlo()
        smtp_obj.login(smtp_user, smtp_password)
        logger.debug(f"send mail {from_addr=}, {to_addrs=}, envelope={msg_mixed.as_string()}") 
        smtp_obj.sendmail(from_addr, (to_addrs+cc_addrs+bcc_addrs), msg_mixed.as_string())
        smtp_obj.quit()
    except Exception as err:
        logger.error(f"Error sending mail", err)        
        raise

def validate_addrs(email):    
    if email is None or not isinstance(email, str) or len(email) == 0:
        logger.warning(f"Invalid email address, removing {email=}")
        return None
    try:        
        stripped = email.strip()
        return validate_email(stripped).email
    except EmailNotValidError as err:        
        logger.error(f"Error parsing email address, {email=}", err)
        return None


def normalize_query(query: str) -> str:
    return re.sub("\s\s+", " ", query).strip()


def execute_query(query):
    query = normalize_query(query)
    logger.info(f"Executing query {query}")
    try:
        return db.execute(query)
    except Exception as err:
        logger.error(f"Error executing {query}", err)
        raise


def set_status(proc, *args):    
    try:
        db.execute(proc, *args)
        logger.info(f"executed {proc=}, args {args=}")
    except Exception as err:
        logger.error(f"Error executing {proc=}, {args=}", err)
        raise


def parse_data(data_obj):
    try:
        data = json.loads(data_obj)        
        if data and isinstance(data, list):
            return dict({"results": data})
        if data and isinstance(data, dict):
            data = data.get("response", data)
            return data
        return None
    except Exception as err:
        logger.error(f"Error parsing data {data_obj=}", err)
        raise


async def request_data(url, **kwargs):
    data_obj = get_data(method="GET", url=url)
    data = parse_data(data_obj)
    if data:
        return kwargs | data
    return kwargs


async def process_query(query, dry_run):
    schema = [
        'exchange_query_id',
        'cli',
        'site_postcode',
        'exchange_name',
        'exchange_code',
        'exchange_postcode',
        'avg_data_usage',
        'implementation_date',
        'file_upload_fk',
        'exchange_query_status_fk'
    ]
    query = { k:query[i] for i, k in enumerate(schema) }

    exchange_query_id = query['exchange_query_id']
    cli = query['cli']
    site_postcode = query['site_postcode']
    avg_data_usage = query['avg_data_usage']
    exchange_code = query['exchange_code']    
    
    if not dry_run:
        args = (exchange_query_id, const.EXCHANGE_QUERY_STATUS_BUSY)
        set_status(const.SP_UPDATE_EXCHANGE_QUERY_STATUS, *args)
    # http://localhost/api/v1.0/pangea/product/pricing/result/13?limit=3.0
    url = f"http://{docker_server_name}:{docker_server_port}/api/v1.0/pangea/product/pricing/result/store/{exchange_query_id}?limit={avg_data_usage}"
    t1 = request_data(url, **query)
    
    criterion = cli if cli else site_postcode if site_postcode else exchange_code
    url = f"http://{docker_server_name}:{docker_server_port}/api/v1.0/pangea/sam/exchange/info?query={criterion}"
    t2 = request_data(url, **query)

    # prop and await results
    results = await asyncio.gather(t1, t2)
    
    # combine dicts using reduce on merge lambda
    data = reduce(lambda x, y: x | y, results, {})

    exchange_code = data['exchange_code']
    if exchange_code:
        url = f"http://{docker_server_name}:{docker_server_port}/api/v1.0/pangea/decommission/exchange/search?query={exchange_code}"
        data = await request_data(url, **data)

        implementation_date = data['implementation_date']
        
        if implementation_date is None:
            url = f"http://{docker_server_name}:{docker_server_port}/api/v1.0/pangea/sam/exchange/info?exchange_code={exchange_code}"
            data = await request_data(url, **data)
    
    implementation_date = data.get('implementation_date')

    if implementation_date:
        try:
            data['stop_sell_date'] = parser.parse(implementation_date).date().strftime("""%Y-%m-%d""")
        except parser.ParserError as err:
            logger.error(f"Unable to parse, {implementation_date=}", err)            
            pass
    
    data = {k:v for k,v in data.items() if v is not None}

    args = (
        data.get('exchange_query_id'),
        data.get('cli', const.NOT_PROVIDED),
        data.get('site_postcode', const.NOT_PROVIDED),
        data.get('exchange_name', const.NO_DATA_RESULTS_FOUND),
        data.get('exchange_code', const.NO_DATA_RESULTS_FOUND),
        data.get('exchange_postcode', const.NO_DATA_RESULTS_FOUND),
        data.get('avg_data_usage'),
        data.get('stop_sell_date', const.NO_STOP_SELL_INFORMATION),        
        data.get('file_upload_fk'),
        const.EXCHANGE_QUERY_STATUS_DONE
    )
    
    proc = const.SP_UPDATE_EXCHANGE_QUERY

    try:
        affected, _ = db.execute(proc, *args)
        logger.info(f"Executed {proc=}, {args=}, {affected=}")        
    except Exception as err:
        logger.error(f"Error updating, {proc=}, {args=}", err)
        if not dry_run:            
            args = (exchange_query_id, const.EXCHANGE_QUERY_STATUS_EXCEPTION)
            set_status(const.SP_UPDATE_EXCHANGE_QUERY_STATUS, *args)
        raise


async def process_upload(upload, dry_run):
    file_upload_id, *_ = upload

    if not dry_run:
        args = (file_upload_id, const.FILE_STATUS_UPLOAD)
        set_status(const.SP_UPDATE_FILE_UPLOAD_STATUS, *args)

    proc = const.SP_GET_EXCHANGE_QUERY_BY_FILE_UPLOAD_ID
    if not dry_run:
        try:
            queries, _ = db.execute(proc, file_upload_id)
            logger.info(
                f"fetched builk queries, {proc=}, {file_upload_id=}, {dry_run=}"
            )
        except Exception as err:
            logger.error(f"Error uploading, {proc=}, raising err", err)
            if not queries and not dry_run:
                args = (file_upload_id, const.FILE_STATUS_EXCEPTION)
                set_status(const.SP_UPDATE_FILE_UPLOAD_STATUS, *args)
            raise

    if not dry_run:
        args = (file_upload_id, const.FILE_STATUS_PROCESS)
        set_status(const.SP_UPDATE_FILE_UPLOAD_STATUS, *args)

    for query in queries:
        logger.info(f"processing {query=}")
        try:
            await process_query(query, dry_run)
        except Exception as err:
            logger.error(f"Error processing upload query, {query=}", err)

    if not dry_run:
        args = (file_upload_id, const.FILE_STATUS_NOTIFY)
        set_status(const.SP_UPDATE_FILE_UPLOAD_STATUS, *args)


async def process_uploads(dry_run):
    try:
        if not dry_run:
            proc, args = const.SP_GET_UPLOADED_FILES, const.FILE_STATUS_NEW
            uploads, _ = db.execute(proc, args)
            for upload in uploads:
                await process_upload(upload, dry_run)
        logger.info(f"fetched uploads, {dry_run=}")
    except Exception as err:
        logger.error(f"Error processing uploads, raising err", err)
        raise


async def process_notification(notification, dry_run):
    file_upload_id, email_to_address, *_ = notification
    try:        
        # http://localhost/api/v1.0/pangea/product/pricing/recommendations/file/upload/3
        url = f"http://{docker_server_name}:{docker_server_port}/api/v1.0/pangea/product/pricing/recommendations/file/upload/{file_upload_id}"
        data = await request_data(url)                    
    except Exception as err:
        logger.error(f"Error executing  {url=}, {data=}, {dry_run=}", err)
        raise
    try:        
        products = []
        email_template_schema
        multi_level_index = [
            email_template_schema['product_class'],
            email_template_schema['product_category'],
            email_template_schema['product_term']
            ]
        single_level_index = [
            email_template_schema['cli'],
            email_template_schema['site_postcode'],
            email_template_schema['exchange_name'],
            email_template_schema['exchange_code'],
            email_template_schema['avg_data_usage'],
            email_template_schema['stop_sell_date'],
            email_template_schema['switch_off_date'],
            email_template_schema['product_limit']
        ]        
        drop_columns = ['exchange_query_status_id', 'exchange_query_id', 'redis_cache_result_key', 'file_email_address', 'file_upload_id']
        for record in data['results']:
            product_df = pandas.DataFrame.from_dict(record['product_pricing'])
            product_df = product_df.drop(['product_name', 'product_unit'], axis=1)
            product_df.rename(columns=email_template_schema, inplace=True)
            prices_df = pandas.DataFrame.from_dict(record)            
            prices_df = prices_df.drop(drop_columns+['product_pricing'], axis=1)
            prices_df.rename(columns=email_template_schema, inplace=True)
            combined_df = pandas.concat([product_df, prices_df], axis=1)
            products.append(combined_df)
        # concat dataframes
        df = pandas.concat(products)
        df[email_template_schema['stop_sell_date']] = pandas.to_datetime(df[email_template_schema['stop_sell_date']])
        df[email_template_schema['switch_off_date']] = pandas.to_datetime(df[email_template_schema['switch_off_date']])        
        df = pandas.pivot_table(df, index=single_level_index, columns=multi_level_index, values=email_template_schema['product_price'])
        df.columns.names = (None, None, None) # reset multi-level index names
        # setup mail
        from_addr = email_from_address
        to_addrs = email_to_address.split(',')
        cc = email_cc_address.split(',')
        bcc = email_bcc_address.split(',')
        # send mail with attachments
        logger.info(f"sending mail {from_addr=}, {to_addrs=}, {cc=}, {bcc=}")
        args = (file_upload_id, const.FILE_STATUS_COMPLETE)
        if not dry_run:
            send_mail(from_addr, to_addrs, cc, bcc, df=df)
            set_status(const.SP_UPDATE_FILE_UPLOAD_STATUS, *args)
        logger.info(f"sent mail {args=}, {df=}")
        if not dry_run:
            args = (file_upload_id, const.FILE_STATUS_SENT)
            set_status(const.SP_UPDATE_FILE_UPLOAD_STATUS, *args)
        logger.info(f"updated status, {args=}")
        return True
    except Exception as err1:
        if not dry_run:
            args = (file_upload_id, const.FILE_STATUS_EXCEPTION)
            set_status(const.SP_UPDATE_FILE_UPLOAD_STATUS, *args)
        logger.error(f"Error sending mail, {notification=}, {email_to_address=}", err1)
        raise


async def process_notifications(dry_run):
    try:
        if not dry_run:
            proc, args = const.SP_GET_UPLOADED_FILES, const.FILE_STATUS_NOTIFY
            notifications, _ = db.execute(proc, args)
            for notification in notifications:
                await process_notification(notification, dry_run)
        logger.info(f"fetched notifications, {dry_run=}")
    except Exception as err:
        logger.error(f"Error processing notification, raising err", err)
        raise


async def _main(**kwargs):
    dry_run = kwargs.get("dry_run", None)
    try:
        await process_uploads(dry_run)
        await process_notifications(dry_run)
    except Exception as ex:
        logger.error(f"Error processing uploads and notifications, {kwargs=}", ex)
        return False
    return True


if __name__ == "__main__":
    """Script args for querying templates and notifications"""
    start = time.time()
    logger.info(f"__main__ call")
    logger.info(f"Started clock {start=}")
    # setup environment, get flask config
    env = os.environ.get("FLASK_ENV", "default")
    config = configs[env]
    # setup app for config keys
    app = create_app(config=config)
    # setup mysql server
    docker_db_name = app.config["DOCKER_DB_NAME"]
    docker_server_name = app.config["DOCKER_SERVER_NAME"]
    docker_server_port = app.config["DOCKER_SERVER_PORT"]    
    docker_proxy_name = app.config["DOCKER_PROXY_NAME"]
    # setup smtp and email
    email_from_address = app.config["EMAIL_FROM_ADDRESS"]
    email_cc_address = app.config["EMAIL_CC_ADDRESSES"]
    email_bcc_address = app.config["EMAIL_BCC_ADDRESSES"]
    email_attachment = app.config["EMAIL_ATTACHMENT"]
    email_subject = app.config["EMAIL_SUBJECT"]
    smtp_user = app.config["SMTP_USER"]
    smtp_password = app.config["SMTP_PASSWORD"]
    smtp_host = app.config["SMTP_HOST"]
    smtp_port = app.config["SMTP_PORT"]
    # setup email template content
    email_template_text = app.config["EMAIL_TEMPLATE_TEXT"]
    email_template_html = app.config["EMAIL_TEMPLATE_HTML"]
    email_template_schema = app.config["EMAIL_TEMPLATE_SCHEMA"]
    # parse args, invoke main
    p = argparse.ArgumentParser(description="Process command line parameters")
    p.add_argument(
        "--dry-run",
        action=argparse.BooleanOptionalAction,
        help="dry run flag, if set to True no writes",
    )
    args = p.parse_args()
    status = asyncio.run(_main(**vars(args)))
    # capture wall time and status
    end = time.time()
    elapsed = end - start
    logger.info(f"Completed, {elapsed=}, {status=}")
